#!/usr/bin/env python3
import os
import glob
import urllib.parse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def build_advanced_excel_report():
    print("[+] Initializing Intelligent Excel Reporter Engine (Summary & SecretFinder Mode)...", flush=True)
    
    # 1. 마스터 타깃 목록 로드
    if not os.path.exists('targets.txt'):
        print("[-] Error: targets.txt missing.", flush=True)
        return
        
    with open('targets.txt', 'r') as f:
        targets = [line.strip() for line in f if line.strip()]

    # 데이터 구조화
    matrix_data = {domain: set() for domain in targets}

    # 2. 12대 가상머신 데이터 전수조사
    txt_files = glob.glob('results/**/*.*', recursive=True) + glob.glob('results/*.*')
    txt_files = [f for f in txt_files if os.path.isfile(f)]

    if not txt_files:
        print("[-] Warning: No decrypted text files found in results/ folder.", flush=True)
        return

    print(f"[+] Processing {len(txt_files)} data source files...", flush=True)
    
    for file_path in txt_files:
        filename = os.path.basename(file_path).lower()
        
        if 'secretfinder' in filename:
            source_tool = 'SecretFinder'
        elif 'waybackurls' in filename:
            source_tool = 'Waybackurls'
        elif 'gau' in filename:
            source_tool = 'GAU'
        else:
            source_tool = 'Combined-Engine'

        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line in f:
                    url = line.strip()
                    if not url or url.startswith('#'):
                        continue
                    
                    for domain in targets:
                        if domain in url:
                            matrix_data[domain].add((url, source_tool))
                            break
        except Exception as e:
            print(f"[-] Error reading {filename}: {e}", flush=True)

    # 3. 엑셀 문서 빌드 시작
    print("[+] Compiling Dashboard & High Risk sheets...", flush=True)
    wb = Workbook()

    font_header = Font(name='Malgun Gothic', size=11, bold=True, color='FFFFFF')
    fill_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')

    # -----------------------------------------------------------------
    # [🔥대시보드 칼럼 재조율] 총 URL 합계와 SecretFinder 건수만 깔끔하게 노출
    # -----------------------------------------------------------------
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    dash_headers = ["No", "Target Domain (대상 도메인)", "Total URLs (총 URL 합계)", "SecretFinder Count (SecretFinder 탐지 건수)"]
    ws_dash.append(dash_headers)
    ws_dash.row_dimensions[1].height = 26
    for col_num, text in enumerate(dash_headers, 1):
        cell = ws_dash.cell(row=1, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    # 2번째 탭: 취약 자산 리스트 기틀 유지
    ws_high = wb.create_sheet(title="High Risk Targets")
    high_headers = ["No", "Domain (도메인)", "High Risk URL / Endpoint (위험 주소)", "Source Tool (탐지 도구)", "Risk Reason (위험 사유)"]
    ws_high.append(high_headers)
    ws_high.row_dimensions[1].height = 26
    for col_num, text in enumerate(high_headers, 1):
        cell = ws_high.cell(row=1, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    high_risk_keywords = ['config', '.env', 'xml', 'json', 'secret', 'api/v', 'token', 'admin', 'password', 'key', 'credential', 'mysql']
    
    dash_idx = 1
    high_risk_idx = 1
    sheets_created = 0

    # 4. 데이터 매핑 및 시트 주입
    print("[+] Injecting pure grid data into Excel sheet matrix...", flush=True)
    for domain, dataset in matrix_data.items():
        if not dataset:
            continue  
            
        # [A] 대시보드 통계 계산 연산
        total_urls = len(dataset) # 도메인별 긁어모은 순수 URL 개수 전체 합산
        secret_criticals = sum(1 for url, tool in dataset if tool == 'SecretFinder') # SecretFinder 발견 개수
        
        ws_dash.append([dash_idx, domain, total_urls, secret_criticals])
        dash_idx += 1

        # [B] High Risk 자산 분류 로직
        sorted_dataset = sorted(list(dataset), key=lambda x: (x[1], x[0]))
        for url, tool in sorted_dataset:
            is_high_risk = False
            reason = ""
            
            if tool == 'SecretFinder':
                is_high_risk = True
                reason = "SecretFinder 소스코드 내 보안 자격증명 노출 의심"
            else:
                url_lower = url.lower()
                matched_keys = [key for key in high_risk_keywords if key in url_lower]
                if matched_keys:
                    is_high_risk = True
                    reason = f"민감 엔드포인트 노출 파라미터 감지 ({', '.join(matched_keys)})"
                    
            if is_high_risk:
                ws_high.append([high_risk_idx, domain, url, tool, reason])
                high_risk_idx += 1

        # [C] 개별 도메인 전용 상세 시트 마감
        safe_tab_name = domain[:30]
        ws = wb.create_sheet(title=safe_tab_name)
        sheets_created += 1

        headers = ["No", "Target URL / Endpoint (수집된 자산 주소)", "Source Tool (발견 도구)"]
        ws.append(headers)
        ws.row_dimensions[1].height = 26
        for col_num in range(1, 4):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center

        # 데이터 고속 사출 루프 (OOM 과부하 원인 완전 배제 구조)
        for idx, (url, tool) in enumerate(sorted_dataset, 1):
            if idx > 1048500:
                break
            ws.append([idx, url, tool])

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 85
        ws.column_dimensions['C'].width = 18

    # 5. 최상단 마스터 탭 2개 레이아웃 가로폭 최종 확정
    ws_dash.column_dimensions['A'].width = 8
    ws_dash.column_dimensions['B'].width = 35
    ws_dash.column_dimensions['C'].width = 25
    ws_dash.column_dimensions['D'].width = 25

    ws_high.column_dimensions['A'].width = 8
    ws_high.column_dimensions['B'].width = 25
    ws_high.column_dimensions['C'].width = 85
    ws_high.column_dimensions['D'].width = 15
    ws_high.column_dimensions['E'].width = 35

    # 6. 마무리 및 마스터 파일 저장
    if sheets_created > 0:
        os.makedirs('reports', exist_ok=True)
        report_path = 'reports/passive_recon_report_v1.xlsx'
        wb.save(report_path)
        print(f"[+] [SUCCESS] Master Report with Summary & SecretFinder Dashboard generated at: {report_path}", flush=True)
    else:
        print("[-] Error: Scan results were empty. Excel file not created.", flush=True)

if __name__ == '__main__':
    build_advanced_excel_report()
